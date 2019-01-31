import openpyxl
file = 'C:/Users/InterTicket/OneDrive/Interticket_anyagok/tesztesetek/2084.xlsx'
file2 = 'C:/Users/InterTicket/OneDrive/Interticket_anyagok/tesztesetek/fizikai_anonimizalas.xlsx'
wb = openpyxl.load_workbook(filename=file)

elsosor = 1
# Melyik sheet-et akarunk használni
ws = wb.worksheets[1]
ws1 = wb.worksheets[0]
szoveg0 = ws.cell(1,1).value

# Hova szeretnék az eredmenyt a ws sheeten ahol az teszmanko van
eredemenyoszlop = 5
eredmenysummary = 5
# Hol ér véget a feldolgozás
sorvege = 71
summary0 ="Teszt során azt nézzük meg, hogy amennyiben lefut a fizikai cron anonimizálás, akkor a megfelelő ügyféladatok azok törlődnek-e."
summary1a = "Hely, adatbázis: "
summary2a = "Adattípus: "
summary3a = "Törölni kell-e: "



#i = ahonnan indul a tesztek feldolgozása
i = 1

for sor in range(i, sorvege):
    oszlop = 1
    szoveg1 = None
    while ws.cell(sor,oszlop+1).value != None:
        oszlop = oszlop + 1
        print(ws.cell(sor,oszlop).value)
        if szoveg1 == None:
            szoveg1 = ws.cell(sor,oszlop).value
        else:
            szoveg1 = szoveg1 + '-' + str(ws.cell(sor, oszlop).value)

        # Summary mezőt összrakjuk
        if oszlop == 2:
            if ws.cell(sor,oszlop).value == "tranzakcios_adatbazis":
                summary1b = "Tranzakciós táblában kell nézni az adatokat"
                action0 = ""
            elif ws.cell(sor,oszlop).value == "Regtrans_adatbazis":
                summary1b = "Regtrans adatbázisban kell nézni az adatokat"
                action0 = " "
        elif oszlop == 3:
            summary2b = ws.cell(sor, oszlop).value
        elif oszlop == 4:
            if ws.cell(sor, oszlop).value == "nem":
                summary3b = "Az adatot nem kell anonimizálni, törölni"
                elvart = "Az adott adat nem törlődött"
            elif ws.cell(sor, oszlop).value == "igen":
                summary3b = "Az adatot anonimizálni, törölni kell"
                elvart = "Az adott adat törlődött"
    print(szoveg0)
    print(szoveg1)

    teljesszoveg = szoveg0 + '-' + szoveg1 + "-story_2084"
    ws.cell(sor, eredemenyoszlop).value = teljesszoveg
    summary1 = summary1a + summary1b
    summary2 = summary2a + summary2b
    summary3 = summary3a + summary3b
    summarylista = []
    summarylista.append(summary0)
    summarylista.append(summary1)
    summarylista.append(summary2)
    summarylista.append(summary3)
    summary = '</br></br>'.join(summarylista)
    #print(ws.cell(sor, eredemenyoszlop).value)
    # ws1.cell(sor, eredmenysummary).value = summary
    elsosor = elsosor + 1
    # teszteset neve ami a fő sheetre kerül
    ws1.cell(elsosor,3).value = teljesszoveg
    # teszteset summary leírása
    ws1.cell(elsosor, 5).value = summary
    action0 = "Megnézzük, hogy az adott adattal mi történt."
    ws1.cell(elsosor,7).value = action0
    ws1.cell(elsosor, 8).value =elvart
wb.save(file2)
wb.close()
