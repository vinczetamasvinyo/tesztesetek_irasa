import openpyxl
file = 'C:/Users/InterTicket/OneDrive/Interticket_anyagok/tesztesetek/story_2013.xlsx'
file2 = 'C:/Users/InterTicket/OneDrive/Interticket_anyagok/tesztesetek/tranzakcio_ido.xlsx'
wb = openpyxl.load_workbook(filename=file)

elsosor = 1
# Melyik sheet-et akarunk használni
ws = wb.worksheets[1]
ws1 = wb.worksheets[0]
szoveg0 = ws.cell(2,1).value

# Hova szeretnék az eredmenyt a ws sheeten ahol az teszmanko van
eredemenyoszlop = 6
eredmenysummary = 5
# Hol ér véget a feldolgozás
sorvege = 50
summary0 ="Teszt során az nézzük, hogy egy vásárlás során a megvásárolt elem/elemek esetében megfelelő ideig " \
          "tároljuk-e az ügyféladatokat. A tárolási idő megfelelően számolódik-e."
summary1a = "Vásárlás típusa: "
summary2a = "Megvásárolt tétel: "
summary3a = "Vásárlások darabszáma: "
summary4a = "Az adott tranzakcióban melyik tétel vásárlási ideje érvényesül, melyik szerint kell a legrégebbi ideig " \
            "tárolni az adatokat:"
summaryutolso1 = "Tárolásra vonatkozó jelenlegi információ az, hogy jegyeknél a tranzakcióban lévő legutolsó előadás " \
                 "időpontja + beállított dátum ami most 210 nap. Bérlet, ajándékkártya, kedvezménykártya esetében 2 éve. Ekkor a tranzakció ideje + 2 évet adunk meg."
summaryutolso2 = "Vegyes kosár(jegy/berlet) esetében pedig meg kell keresni hogy melyik a legrégebbi időpont ameddig " \
                 "tárolni kellene az ügyféladatokat és azt kell beállítani."


#i = ahonnan indul a tesztek feldolgozása
i = 2

for sor in range(i, sorvege):
    oszlop = 1
    szoveg1 = None
    while ws.cell(sor,oszlop+1).value != None:
        oszlop = oszlop + 1
        print(ws.cell(sor,oszlop).value)
        if szoveg1 == None:
            szoveg1 = ws.cell(sor,oszlop).value
        else:
            szoveg1 = szoveg1 + '-' + str(ws.cell(sor, oszlop).value)

        # Summary mezőt összrakjuk
        if oszlop == 2:
            if ws.cell(sor,oszlop).value == "uj_vasarlas":
                summary1b = "Új vásárlást csinálunk"
                action0 = "Elvégezzük a summary esetben leírt vásárlást"
            elif ws.cell(sor,oszlop).value == "migracio_regiek":
                summary1b = "Migrációs egy korábbi vásárláshoz számoljuk ki az adatot"
                action0 = " Egy korábbi vásárlás esetében nézzük meg azt, hogy a tárolási idő kiszámítása megfelelő-e."
        elif oszlop == 3:
            if ws.cell(sor,oszlop).value == "jegyvasarlas":
                summary2b = "Jegyet vásárolunk"
            elif ws.cell(sor,oszlop).value == "berlet":
                summary2b = "Bérletet vásárolunk"
            elif ws.cell(sor,oszlop).value == "ajandekkartya":
                summary2b = "ajándékkártyát vásárolunk"
            elif ws.cell(sor,oszlop).value == "kedvezmenykartya":
                summary2b = "Kedvezménykártyát vásárolunk"
            elif ws.cell(sor, oszlop).value == "jegy_berlet":
                summary2b = "Jegyet és bérletet is vásárolunk"
            elif ws.cell(sor, oszlop).value == "jegy_berlet_ajandekkartya":
                summary2b = "Jegyet, bérletet és ajándékkártyát is vásárolunk."
            elif ws.cell(sor, oszlop).value == "jegy_berlet_ajandekkartya_kedvez":
                summary2b = "Jegyet, bérletet, ajdánékkártyát, és kevezménykártyát is vásárolunk."
        elif oszlop == 4:
            if ws.cell(sor, oszlop).value == "egy":
                summary3b = "Egy darabot vásárolunk"
            elif ws.cell(sor, oszlop).value == "tobb":
                summary3b = "Több darabot vásárolunk a tételekből, amik különböz időpontra szólnak."

        elif oszlop == 5:
            if ws.cell(sor, oszlop).value == "ido_jegy":
                summary4b = "Jegyvásárlás miatt kell a legtovább tárolni a tranzakció adatokat."
                elvart = 'Az adott adat nem jelenik meg.'
            elif ws.cell(sor, oszlop).value == "ido_berlet":
                summary4b = "A bérlet miatt kell a legtovább tárolni az adatokat."
                elvart = 'az adott adat megjelenik'
            elif ws.cell(sor, oszlop).value == "ido_ajandekkartya":
                summary4b = "Az ajándékkártya miatt kell a legtovább tárolni a tranzakcióban lévő adatokat"
            elif ws.cell(sor, oszlop).value == "ido_kedvezmenykartya":
                summary4b = "A kedvezménykártya miatt kell a legtovább tárolni az adatokat."

    print(szoveg0)
    print(szoveg1)

    teljesszoveg = szoveg0 + '-' + szoveg1 + "-story_2013"
    ws.cell(sor, eredemenyoszlop).value = teljesszoveg
    summary1 = summary1a + summary1b
    summary2 = summary2a + summary2b
    summary3 = summary3a + summary3b
    summary4 = summary4a + summary4b
    summarylista = []
    summarylista.append(summary0)
    summarylista.append(summary1)
    summarylista.append(summary2)
    summarylista.append(summary3)
    summarylista.append(summary4)
    summarylista.append(summaryutolso1)
    summarylista.append(summaryutolso2)
    summary = '</br></br>'.join(summarylista)
    #print(ws.cell(sor, eredemenyoszlop).value)
    # ws1.cell(sor, eredmenysummary).value = summary
    elsosor = elsosor + 1
    # teszteset neve ami a fő sheetre kerül
    ws1.cell(elsosor,3).value = teljesszoveg
    # teszteset summary leírása
    ws1.cell(elsosor, 5).value = summary
    elvart = "Az adott tranzakció esetében az adott logika alapján számolódik ki az adattárolási idő."

    ws1.cell(elsosor,7).value = action0
    ws1.cell(elsosor, 8).value =elvart
wb.save(file2)
wb.close()
