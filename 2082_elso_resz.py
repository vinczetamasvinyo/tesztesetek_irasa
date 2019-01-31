import openpyxl
file = 'C:/Users/InterTicket/OneDrive/Interticket_anyagok/tesztesetek/2082_elso_resz.xlsx'
file2 = 'C:/Users/InterTicket/OneDrive/Interticket_anyagok/tesztesetek/2082_elso_reszc.xlsx'
wb = openpyxl.load_workbook(filename=file)

elsosor = 1
# Melyik sheet-et akarunk használni
ws = wb.worksheets[1]
ws1 = wb.worksheets[0]
szoveg0 = ws.cell(1,1).value

# Hova szeretnék az eredmenyt a ws sheeten ahol az teszmanko van
eredemenyoszlop = 5
eredmenysummary = 5
# Hol ér véget a feldolgozás
sorvege = 105
summary0 ="Teszt során azt nézzük meg, hogy amennyiben lejárt egy tranzakciónak a tárolási ideje és IT userrel a megadott helyen rákeresünk akkor nem szabad eredményt felhoznia, illetve ha egyéb adatok alapján megtaláljuk a tranzakciót abban nem szabad személyes adatot mutatni."
summary1a = "Hely az adminban:"
summary2a = "Adat/adattípus: "
summary3a = "Megjelenik-e:"

#i = ahonnan indul a tesztek feldolgozása
i = 1

for sor in range(i, sorvege):
    oszlop = 1
    szoveg1 = None
    while ws.cell(sor,oszlop+1).value != None:
        oszlop = oszlop + 1
        print(ws.cell(sor,oszlop).value)
        if szoveg1 == None:
            szoveg1 = ws.cell(sor,oszlop).value
        else:
            szoveg1 = szoveg1 + '-' + str(ws.cell(sor, oszlop).value)

        # Summary mezőt összrakjuk
        if oszlop == 2:
            if ws.cell(sor,oszlop).value == "Tabla":
                summary1b = "Adminban található táblázatban nézzük az adatokat."
                action0 = "Az adott adatot megnézzük, hogy a táblázatban látható-e."
            elif ws.cell(sor,oszlop).value == "Kereso":
                summary1b = "A kereső felüelten keresünk rá az adatra."
                action0 = "Az adott adatra keresünk a keresőben(2 helyen is van)."
            elif ws.cell(sor,oszlop).value == "tr_reszletek":
                summary1b = "Tranzakció részleteiben nézzük meg az adatokat."
                action0 = "A tranzakció részletekben nézzük meg, hogy megjelenik-e az adat"
            elif ws.cell(sor,oszlop).value == "Pdf_fajl":
                summary1b = "Tranzakción belül a PDF fájl résznél keresünk adatokat."
                action0 = "A PDF fájlban nézzük, hogy megjelenik-e az adat."
            elif ws.cell(sor, oszlop).value == "vissza_ig":
                summary1b = "Tranzakción belül a visszaigazoló email részt nézzük."
                action0 = "A visszaigazoló email részben nézzük, hogy megjelenik-e az adat."
            elif ws.cell(sor, oszlop).value == "export_xls":
                summary1b = "Az excelbe történő exportálásban nézzük meg az adatot"
                action0 = "Az excelbe történő exportálási részben nézzük meg, hogy megjelenik-e az adat"
            elif ws.cell(sor, oszlop).value == "export_csv":
                summary1b = "A csv fájlba történő exportálás során ézzük meg az adatokat."
                action0 = "A CSV történő exportálási részben nézzük meg, hogy adat megjelenik-e."
        elif oszlop == 3:
            summary2b = ws.cell(sor, oszlop).value
        elif oszlop == 4:
            if ws.cell(sor, oszlop).value == "nincs":
                summary3b = "Az adatnak nem szabad megjelennie, illetve ha erre keresünk akkor az nem hozhat eredményt."
                elvart = "Az adat nem jelenthet meg. Vagy ha keresésről szól a teszt, akkor az nem vezethet " \
                         "eredményhez"
            elif ws.cell(sor, oszlop).value == "van":
                summary3b = "Az adat megjelenhet. Illetve ha erre az adatra keresünk, akkor megtaálhatjuk a tranzakciót"
                elvart = "Az adat megjelenhet. "

    print(szoveg0)
    print(szoveg1)

    teljesszoveg = szoveg0 + '-' + szoveg1 + "-story_2082_elso"
    ws.cell(sor, eredemenyoszlop).value = teljesszoveg
    summary1 = summary1a + summary1b
    summary2 = summary2a + summary2b
    summary3 = summary3a + summary3b
    summarylista = []
    summarylista.append(summary0)
    summarylista.append(summary1)
    summarylista.append(summary2)
    summarylista.append(summary3)
    summary = '</br></br>'.join(summarylista)
    #print(ws.cell(sor, eredemenyoszlop).value)
    # ws1.cell(sor, eredmenysummary).value = summary
    elsosor = elsosor + 1
    # teszteset neve ami a fő sheetre kerül
    ws1.cell(elsosor,3).value = teljesszoveg
    # teszteset summary leírása
    ws1.cell(elsosor, 5).value = summary
    ws1.cell(elsosor,7).value = action0
    ws1.cell(elsosor, 8).value =elvart
wb.save(file2)
wb.close()
