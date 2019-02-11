import openpyxl
# itt határozzuk meg, hogy melyik fájlból dolgozunk. Ez az alap fájl.
file = 'C:/Users/InterTicket/OneDrive/Interticket_anyagok/tesztesetek/kimutatas_ajandekkartyakrol_folyamatban3.xlsx'
# ebbe a fájlba mentjük el a elkészített módosításokat.
file2 = 'C:/Users/InterTicket/OneDrive/Interticket_anyagok/tesztesetek/kimutatas_ajandekkartyakrol_folyamatban3_mentes.xlsx'
# Megnyitjuk az alap fájlt.
wb = openpyxl.load_workbook(filename=file)

elsosor = 1
# Melyik sheet-et akarunk használni a file egy esetében
ws = wb.worksheets[1]
ws1 = wb.worksheets[0]
szoveg0 = ws.cell(1,1).value

# Hova szeretnék az eredmenyt a ws sheeten ahol az teszmanko a tesztvazlatok vannak.
eredemenyoszlop = 6
eredmenysummary = 6
# Hol ér véget a feldolgozás
sorvege = 259
# Az első alap szöveg
summary0 ="Teszt során azt nézzük meg, hogy amennyiben a megadott paraméterekkel futtatjuk kimutatás az ajándékkártya riportot, akkor minden rendben működik-e."
summary1a = "Első opció: "
summary2a = "További beállítási opció az első opción belül:"
summary3a = "A keresési feltételnek szerepelnie kell-e a riportban:"
summary4a = "Mibe kérjük le a riportot:"

elvart_eredmeny = ""



#i = ahonnan indul a tesztek feldolgozása
i = 2

for sor in range(i, sorvege):
    elvart = ""
    elvart_eredmeny0 = ""
    oszlop = 1
    szoveg1 = None
    while ws.cell(sor,oszlop+1).value != None:
        oszlop = oszlop + 1
        print(ws.cell(sor,oszlop).value)
        # itt rakjuk össze a teszteset nevét
        if szoveg1 == None:
            szoveg1 = ws.cell(sor,oszlop).value
        else:
            szoveg1 = szoveg1 + '-' + str(ws.cell(sor, oszlop).value)

        # Summary mezőt összrakjuk
        if oszlop == 2:
            if ws.cell(sor,oszlop).value == "ajandekkartya_tipus":
                summary1b = "Az ajándékkártya típusa lehetőségeit fogjuk állítani"
                action0 = ""
            elif ws.cell(sor,oszlop).value == "ajandekkartya_neve":
                summary1b = "Az ajándékkártya neve blockban történik majd a beállítás"
                action0 = " "
            elif ws.cell(sor,oszlop).value == "Vonalkod":
                summary1b = "A vonalkód mezőbe fogunk egy vonalkódot beírni."
            elif ws.cell(sor,oszlop).value == "Vasarlas_datuma":
                summary1b = "A vásárlás dátuma beállításokon belül fogjuk állítani a dolgokat"
            elif ws.cell(sor, oszlop).value == "Bevaltas_idopontja":
                summary1b = "A vásárlás dátuma beállításokon belül fogjuk állítani a dolgokat"
            elif ws.cell(sor, oszlop).value == "Bevaltas_idopontja":
                summary1b = "A Beváltás időpontja beállításokon belül fogjuk állítani a dolgokat"
            elif ws.cell(sor, oszlop).value == "vasarlas_bevaltas_idopontja":
                summary1b = "A vásárlás és a beváltás időpontja beállításokon belül fogjuk állítani a dolgokat"
            elif ws.cell(sor, oszlop).value == "vasarlas_bevaltas_idopontja":
                summary1b = "A felhasználói részen belül állítjuk a lehetőségeket"
            elif ws.cell(sor, oszlop).value == "Ugyfel":
                summary1b = "Az ügyfél választási résznél fogjuk állítani a lehetőségeket"
            elif ws.cell(sor, oszlop).value == "Ajandek_tipusa":
                summary1b = "Azt fogjuk nézni, hogy az ajándékkártya hálózatos, helyi, vagy mindkettő."
            elif ws.cell(sor, oszlop).value == "Status":
                summary1b = "Az ajándékkártya státusza részen belül fogjuk állítani a lehetőségeket"
            elif ws.cell(sor, oszlop).value == "Hol_ertekesitett":
                summary1b = "Azt fogjuk állítani, hogy hol lett értékesítve az ajándékkártya, hálózaton, helyi vagy mindkettőben."
            elif ws.cell(sor, oszlop).value == "Mind_kibontva":
                summary1b = "A Mind kibotva checkbox-t fogjuk majd állítani."
        elif oszlop == 3:
            if ws.cell(sor, oszlop).value == "egycelu":
                summary2b = "Egycélú ajándékkártya lesz kiválasztva"
            elif ws.cell(sor, oszlop).value == "tobbcelu":
                summary2b = "többcélú ajándékkártya lesz kiválasztva"
            elif ws.cell(sor, oszlop).value == "osszes_ajandekkartya":
                summary2b = "Összes ajándékkártya lesz kiválasztva"
            elif ws.cell(sor, oszlop).value == "lista_egycelu":
                summary2b = "A listán belül egy egycélú utalványt válaztunk ki"
            elif ws.cell(sor, oszlop).value == "lista_tobbcelu":
                summary2b = "A listán belül egy többcélú utalványt válaztunk ki"
            elif ws.cell(sor, oszlop).value == "osszes_ajandekkartya":
                summary2b = "Az összes ajándékkártya opciót használjuk"
            else:
                summary2b = ws.cell(sor, oszlop).value
        elif oszlop == 4:
            if ws.cell(sor, oszlop).value == "van":
                summary3b = "Pozitív teszt olyan ajándékkártya szükséges ami biztos, hogy szerepel majd a riportban."
                elvart = "A megadott feltételeknek megfelelő ajándékutalvány megjelenik a riportban"
            elif ws.cell(sor, oszlop).value == "nincs":
                summary3b = "Negatív teszt olyan ajándékkártya szükséges ami biztos, hogy nem szerepel majd a riportban"
                elvart = "A megadott feltételeknek megfelelő ajándékutalvány NEM jelenik meg a riportban."
        elif oszlop == 5:
            if ws.cell(sor, oszlop).value == "Kepernyore":
                summary4b = "Képernyőre kérjük le a riportot"
            elif ws.cell(sor, oszlop).value == "Nyomtatora":
                summary4b = "Nyomtatóra kérjük le a riportot"
            elif ws.cell(sor, oszlop).value == "Excel":
                summary4b = "Excel fájlba kérjük le a riport"
    print(szoveg0)
    print(szoveg1)

    teljesszoveg = szoveg0 + '-' + szoveg1 + "-story_valami"
    ws.cell(sor, eredemenyoszlop).value = teljesszoveg
    summary1 = summary1a + summary1b
    summary2 = summary2a + summary2b
    summary3 = summary3a + summary3b
    summary4 = summary4a + summary4b
    summarylista = []
    summarylista.append(summary0)
    summarylista.append(summary1)
    summarylista.append(summary2)
    summarylista.append(summary3)
    summarylista.append(summary4)
    summary = '</br></br>'.join(summarylista)
    #print(ws.cell(sor, eredemenyoszlop).value)
    # ws1.cell(sor, eredmenysummary).value = summary
    elsosor = elsosor + 1
    # teszteset neve ami a fő sheetre kerül
    ws1.cell(elsosor,3).value = teljesszoveg
    # teszteset summary leírása
    ws1.cell(elsosor, 5).value = summary
    action0 = "Megnézzük, hogy a summary-ben leírt módon működik-e a riport."
    ws1.cell(elsosor,7).value = action0
    ws1.cell(elsosor, 8).value =elvart
wb.save(file2)
wb.close()
