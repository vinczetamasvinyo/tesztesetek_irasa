import openpyxl
file = 'C:/Users/InterTicket/OneDrive/Interticket_anyagok/tesztesetek/story_2013.xlsx'
file2 = 'C:/Users/InterTicket/OneDrive/Interticket_anyagok/tesztesetek/tranzakcio_ido.xlsx'
wb = openpyxl.load_workbook(filename=file)

elsosor = 1
# Melyik sheet-et akarunk használni
ws = wb.worksheets[1]
ws1 = wb.worksheets[0]
szoveg0 = ws.cell(1,2).value

# Hova szeretnék az eredmenyt
eredemenyoszlop = 6
eredmenysummary = 5
# Hol ér véget a feldolgozás
sorvege = 209
summary0 ="A teszt során azt nézzük, hogy a megadott user típussal az adott helyen az adott adattípus megfelelően jelenik-e meg."
summary1a = "User típus: "
summary2a = "Hely:  "
summary3a = "Adattípus:"
summary4a = "Meg kell-e jelennie:"

#i = ahonnan indul a tesztek feldolgozása
i = 1

for sor in range(i, sorvege):
    oszlop = 1
    szoveg1 = None
    while ws.cell(sor,oszlop+1).value != None:
        oszlop = oszlop + 1

        if szoveg1 == None:
            szoveg1 = ws.cell(sor,oszlop).value
        else:
            szoveg1 = szoveg1 + '-' + str(ws.cell(sor, oszlop).value)

        # Summary mezőt összrakjuk
        if oszlop == 2:
            if ws.cell(sor,oszlop).value == "Partner":
                summary1b = " Partner felhasználó"

            elif ws.cell(sor,oszlop).value == "IT":
                summary1b = "Interticket felhasználó"
        elif oszlop == 3:
            if ws.cell(sor,oszlop).value == "Tabla":
                summary2b = "Admin alján közepén található táblázat"
            elif ws.cell(sor,oszlop).value == "Kereso":
                summary2b = "Admin tetején található kereső felület"
            elif ws.cell(sor,oszlop).value == "tr_reszletek":
                summary2b = "Adott tranzakció amit megnyitunk."
            elif ws.cell(sor,oszlop).value == "Pdf_fajl":
                summary2b = "A tranzakció-t ha megnyitjuk, akkor az alján található pdf fájl rész"
            elif ws.cell(sor, oszlop).value == "vissza_ig":
                summary2b = "A tranzakció-t ha megnyitjuk, akkor az alján található visszaigazoló email rész"
            elif ws.cell(sor, oszlop).value == "export_xls":
                summary2b = "XLS-be történő exportálás"
            elif ws.cell(sor, oszlop).value == "export_csv":
                summary2b = "CSV-be történő exportálás"
        elif oszlop == 4:
            if ws.cell(sor, oszlop).value == "nev":
                summary3b = 'Név'
            elif ws.cell(sor, oszlop).value == "email":
                summary3b = "E-mail cím"
            elif ws.cell(sor, oszlop).value == "Kereses":
                summary3b = "Kereső mező"
            elif ws.cell(sor, oszlop).value == "tr_id":
                summary3b = "Tranzakció ID"
            elif ws.cell(sor, oszlop).value == "datum":
                summary3b = "Dátum"
            elif ws.cell(sor, oszlop).value == "osszeg_darab":
                summary3b ="Összeg / darabszám"
            elif ws.cell(sor, oszlop).value == "kez_mod":
                summary3b = "Kézbesítési mód"
            elif ws.cell(sor, oszlop).value == "fiz_mod":
                summary3b = "Fizetési mód"
            elif ws.cell(sor, oszlop).value == "utalvany":
                summary3b = "utalvány"
                # summary3b = ws.cell(sor, oszlop).value
            elif ws.cell(sor, oszlop).value == "affiliate":
                summary3b = "Affiliate kód"
            elif ws.cell(sor, oszlop).value == "xls":
                summary3b = "XLs export gomb"
            elif ws.cell(sor, oszlop).value == "csv":
                summary3b = "csv export gomb"
            elif ws.cell(sor, oszlop).value == "br_id":
                summary3b = "Browser id"
            else:
                summary3b = ws.cell(sor, oszlop).value
        elif oszlop == 5:
            if ws.cell(sor, oszlop).value == "nincs":
                summary4b = "Az adott adatnak nem kell megjelennie"
                elvart = 'Az adott adat nem jelenik meg.'
            elif ws.cell(sor, oszlop).value == "van":
                summary4b = "Az adott adatnak meg kell jelennie."
                elvart = 'az adott adat megjelenik'


    if oszlop != eredemenyoszlop:
        print(szoveg0)
        print(szoveg1)

        teljesszoveg = szoveg0 + '-' + szoveg1
        ws.cell(sor, eredemenyoszlop).value = teljesszoveg
        summary1 = summary1a + summary1b
        summary2 = summary2a + summary2b
        summary3 = summary3a + summary3b
        summary4 = summary4a + summary4b
        summarylista = []
        summarylista.append(summary0)
        summarylista.append(summary1)
        summarylista.append(summary2)
        summarylista.append(summary3)
        summarylista.append(summary4)
        summary = '</br></br>'.join(summarylista)
        #print(ws.cell(sor, eredemenyoszlop).value)
        # ws1.cell(sor, eredmenysummary).value = summary
        elsosor = elsosor + 1
        # teszteset neve
        ws1.cell(elsosor,3).value = teljesszoveg
        # teszteset summary leírása
        ws1.cell(elsosor, 5).value = summary
        action0 = "Belépünk " +ws.cell(sor,2).value \
                  + "userrel és megnézzük, hogy az adott helyen az adtok elérhetőek-e."

        action = action0 +", " + summary2 +", " + summary3
        ws1.cell(elsosor,7).value = action
        ws1.cell(elsosor, 8).value =elvart
wb.save(file2)
wb.close()



