import openpyxl
# itt határozzuk meg, hogy melyik fájlból dolgozunk. Ez az alap fájl.
file = 'C:/Users/InterTicket/OneDrive/Interticket_anyagok/tesztesetek/box_office_553.xlsx'
# ebbe a fájlba mentjük el a elkészített módosításokat.
file2 = 'C:/Users/InterTicket/OneDrive/Interticket_anyagok/tesztesetek/box_office_553_2.xlsx'
# Megnyitjuk az alap fájlt.
wb = openpyxl.load_workbook(filename=file)

elsosor = 1
# Melyik sheet-et akarunk használni a file egy esetében
ws = wb.worksheets[1]
ws1 = wb.worksheets[0]
szoveg0 = ws.cell(1,1).value

# Hova szeretnék az eredmenyt a ws sheeten ahol az teszmanko a tesztvazlatok vannak.
eredemenyoszlop = 6
eredmenysummary = 6
# Hol ér véget a feldolgozás
sorvege = 321
# Az első alap szöveg
summary0 ="Teszt során azt nézzük, hogy az aktuális funkció a jegypénztár/Eladáson belül jól működik-e."
summary1a = "Teszt típusa: "
summary2a = "Amit nézünk az oldalon: "
summary3a = "Nézett funkció működése, rövid leírása: "
summary4a = "Használt Böngésző: "
summary5a = "Nyelv: "

elvart_eredmeny = ""
"""
alapteszt1a = "Be kell lép az aktuális városkártya oldalon."
alapteszt1b = "Belépés sikerüt"
alapteszt2a = "Be kell menni a Jegypénztár/Eladás menüpontba"
alapteszt2b = "A belépés sikerült és betöltődéskor látszódnak az események."
"""

#i = ahonnan indul a tesztek feldolgozása
i = 1

for sor in range(i, sorvege):
    elvart = ""
    elvart_eredmeny0 = ""
    oszlop = 1
    szoveg1 = None
    summary1b = ""
    summary2b = ""
    summary3b = ""
    summary4b = ""
    summary5b = ""
    while ws.cell(sor,oszlop+1).value != None:
        oszlop = oszlop + 1
        print(sor)
        print(ws.cell(sor,oszlop).value)
        # itt rakjuk össze a teszteset nevét
        if szoveg1 == None:
            szoveg1 = ws.cell(sor,oszlop).value
        else:
            szoveg1 = szoveg1 + '-' + str(ws.cell(sor, oszlop).value)

        # Summary mezőt összrakjuk
        if oszlop == 2:
            if ws.cell(sor,oszlop).value == "funkcio":
                summary1b = "Funkcionális tesztet csinálunk."
                action0 = ""
            elif ws.cell(sor,oszlop).value == "ajandekkartya_neve":
                summary1b = "Design böngészőhöz kapcsolódó tesztet csinálunk."
                action0 = " "
        elif oszlop == 3:
            if ws.cell(sor, oszlop).value == "megjelenitett_adat":
                summary2b = "Az Event listában nézzük, hogy az ott szükséges adatok megjelennek-e"
            elif ws.cell(sor, oszlop).value == "megjelenitett_fejlec_szoveg":
                summary2b = "Az Event lista(táblázat) fölött nézzük, hogy a fejléc szövege megjelenik-e."
            elif ws.cell(sor, oszlop).value == "paging":
                summary2b = "A lap alján található listbox-t nézzük amivel be lehet állítani, hogy mennyi adat jelenjen meg az oldalon"
            elif ws.cell(sor, oszlop).value == "nyilak":
                summary2b = "A lap aljánt található lapozó nyilak funkcióját teszteljük"
            elif ws.cell(sor, oszlop).value == "lista_rendezese":
                summary2b = "A lapon megjelenő események listáját nézzük, hogy azok megfelelően vannak-e rendezve"
            elif ws.cell(sor, oszlop).value == "responsive":
                summary2b = "Responsive működést nézzük meg."
            else:
                summary2b = ws.cell(sor, oszlop).value
        elif oszlop == 4:
            if ws.cell(sor, oszlop).value == "program_neve":
                summary3b = "Azt nézzük meg, hogy a program neve megjelenik-e az adatok közt."
                action0 =" "
                elvart = "A program neve megjelenik az események listájában"
            elif ws.cell(sor, oszlop).value == "helyszin_neve":
                summary3b = "Azt nézzük meg, hogy a helyszín neve megjelenik-e az adatok közt."
                elvart = "A helyszín neve pl:MÜPA megjelenik az adatok közt."
            elif ws.cell(sor, oszlop).value == "auditorium_neve":
                summary3b = "Azt nézzük meg, hogy az auditorium(nézőtér) neve megjelenik az adatok közt."
                elvart = "Az auditorium, nézőtér neve megjelent az event listákban"
            elif ws.cell(sor, oszlop).value == "eloadas_datuma":
                summary3b = "Azt nézzük meg, hogy az előadás dátuma megjelenik-e."
                elvart ="Az előadás dátuma megjelent az event listában"
            elif ws.cell(sor, oszlop).value == "aladott_jegyek_szama":
                summary3b = "Azt nézzük meg, hogy az eladott jegyek száma megjelenik-e"
                elvart = "Azt nézzük meg, hogy az eladott jegyek száma megjelent-e."
            elif ws.cell(sor, oszlop).value == "elerheto_jegyek_szama":
                summary3b = "Azt nézzük meg, hogy az elérhető jegyek száma megjelenik-e."
                elvart = "Azt nézzük meg, hgoy az elérhető jegyek száma megjelenik-e."
            elif ws.cell(sor, oszlop).value == "megyjegyzes_mezo":
                summary3b = "Azt nézzük meg, hogy a megjegyzés mező megjelenik-e."
                elvart = "At nézzük meg, hogy a megyjezés mező megjelenik-e az event listában."
            elif ws.cell(sor, oszlop).value == "megyjegyzes_mezo_kissebb_100":
                summary3b = "Azt nézzük meg, hogy a megjegyzés mező rendben megjelenik-e, ha kb 100 karakter hosszú szöveg van benne"
                elvart = "A teljes szöveg rendben megjelenik"
            elif ws.cell(sor, oszlop).value == "megyjegyzes_mezo_kb_280":
                summary3b = "Azt nézzük meg, hogy a megyjegyzés mezőbe egy kb 280 hosszú szöveg van, akkor az rendben megjelenik-e."
                elvart = "A teljes szöveg rendben megjelenik"
            elif ws.cell(sor, oszlop).value == "megyjegyzes_mezo_kb_280":
                summary3b = "Azt nézzük meg, hogy a megjegyzés mezőbe egy több mint 280 hosszú szöveg kerül és akkor mi történik."
                elvart = "A teljes szöveg rendben megjelenik"
            elif ws.cell(sor, oszlop).value == "eloadas":
                summary3b = "Azt nézzük meg, hogy az Előadás szöveg megjelenik-e a táblázat fölött."
                elvart = "A táblázat fölött megjelent az Előadás szöveg"
            elif ws.cell(sor, oszlop).value == "datum":
                summary3b = "Azt nézzük meg, hogy a Dátum szöveg megjelenik-e a táblázat fölött."
                elvart = "A táblázat fölött megjelent az dátum szöveg."
            elif ws.cell(sor, oszlop).value == "jegyek":
                summary3b = "Azt nézzük meg, hogy a jegyek szöveg megjelenik-e a táblázat fölött."
                elvart = "A táblázat fölött megjelent a jegyek szöveg"
            elif ws.cell(sor, oszlop).value == "Megjegyzes":
                summary3b = "Azt nézzük meg, hogy a Megjegyzés szöveg megjelenik-e a táblázato fölött"
                elvart = "A táblázat fölött megjelent Megjegyzés szöveg"
            elif ws.cell(sor, oszlop).value == "lista_elemeinek_megnezese":
                summary3b = "Azt nézzük meg, hogy a lapozóban lévő listaelemek(10,25, 50, 100) megfelelően jelennek-e meg."
                elvart = "A lapozó listában csak a 10, 25, 50, 100-as lista jelenhet meg"
            elif ws.cell(sor, oszlop).value == "10_es_lista_mukodese":
                summary3b = "Az nézzük meg, hogy a 10-es megjelenítési lista jól működik-e."
                elvart = "Maximum 10 event lista elem jelenhet meg. Kevesebb akkor jelenik meg, ha kevesebb lista van mint 10."
            elif ws.cell(sor, oszlop).value == "25_lista_mukodese":
                summary3b = "Azt nézzük meg, hogy amennyiben 25 esemény lista megjelenítését választjuk, akkor ez jól jelenik-e meg."
                elvart = "Maximum 25 event lista elem jelenhet meg. Kevesebb akkor jelenik meg, ha kevesebb lista van mint 25."
            elif ws.cell(sor, oszlop).value == "50_lista_mukodese":
                summary3b = "Azt nézzük meg, hogy amennyiben 50 esemény lista megjelenítését választjuk, akkor ez jól jelenik-e meg."
                elvart = "Maximum 50 event lista elem jelenhet meg. Kevesebb akkor jelenik meg, ha kevesebb lista elem van mint 50."
            elif ws.cell(sor, oszlop).value == "100_lista_mukodese":
                summary3b = "Azt nézzük meg, hogy amennyiben 100 esemény lista megjelenítését választjuk, akkor ez jól jelenik-e meg."
                elvart = "Maximum 100 event lista elem jelenhet meg. Kevesebb akkor jelenik meg, ha kevesebb lista elem van mint 100."
            elif ws.cell(sor, oszlop).value == "legelso_oldal_inaktiv":
                summary3b = "Az nézzük meg, hogyha a legelső oldalon vagyunk és megnyomjuk az ugrás az első oldalra gomb "
                elvart = "A gomb státusza inaktív hisz a legelső oldalon vagyunk, illetve nem tudunk rákattintani."
            elif ws.cell(sor, oszlop).value == "legelso_oldal_aktiv":
                summary3b = "Az nézzük meg, hogyha a legelső oldalról elnavigálunk és megnyomjuk az első oldalra történő ugrást."
                elvart = "A gomb státusza aktív, hisz nem a legelső oldalon vagyunk. Ha megynyomjuk a gombot, akkor a legelső oldalra jutunk"
            elif ws.cell(sor, oszlop).value == "egyel_vissza_inaktiv":
                summary3b = "Megnézzük, hogy az eggyel vissza gomb státusza inaktív-e."
                elvart = "A gomb státusza inaktív hisz az első oldalon vagyunk és onnan nem mehetünk előre"
            elif ws.cell(sor, oszlop).value == "egyel_vissza_aktiv":
                summary3b = "Ha elnavigálunk az első oldalról akkor megnézzük az egyel vissza gomb státuszát és funkcióját."
                elvart = "A gomb státusza aktív hisz már nem az első oldalon vagyunk. Ha megnyomjuk a gombot, akkor egyel előre lapozunk."
            elif ws.cell(sor, oszlop).value == "egyel_elore_aktiv":
                summary3b = "Megnézzük azt, hogyha nem az utolsó oldalon vagyunk, hogy az egyel előre gombt státusza aktív-e, illetve hogy előre lapoz-e."
                elvart = "A gomb státusza aktív hisz előre tudunk lapozni. Ha megnyomjuk, akkor egyel előre lapoz."
            elif ws.cell(sor, oszlop).value == "egyel_elore_inaktiv":
                summary3b = "Megnézzük azt, hogyha a legutolsó oldalra navigálunk akkor az egyel előre gomb inaktív lesz-e."
                elvart = "A gomb státusza inaktív és ha rákattintunk, akkor nem visz előre."
            elif ws.cell(sor, oszlop).value == "utolso_oldal_aktiv":
                summary3b = "Megnézzük azt, hogyha nem az utolsó oldalon vagyunk, akkor az utolsó oldalra navigál gombt aktív-e, illetve ha rákattintunk mit történik"
                elvart = "A gomb aktív. Ha rákattintunk, akkor elnavigál az utolsó oldalra"
            elif ws.cell(sor, oszlop).value == "utolso_oldal_inaktiv":
                summary3b = "Megnézzük, hogyha az utolsó oldalon vagyunk, akkor milyen az utolsó oldalra navigál gombt státusza, illetve ha rákattintunk mi történik."
                elvart = "A gomb inaktív és nem lehet rákattintani."
            elif ws.cell(sor, oszlop).value == "datum_szerint_novekvo":
                summary3b = "Megnézzük, hogy dátum szerint növekvő sorrendben vannak-e az események"
                elvart = "Dátum szerint növekvő sorrendben vannak az események"
            elif ws.cell(sor, oszlop).value == "azonos_datum_abc":
                summary3b = "Megnézzük, hogy dátum szerint, majd abc szerint növekvő sorrendben vannak-e az események."
                elvart = "Elsőnek dátum szerint vannak növekvő sorrendben az események. Ha azonos dátumon több eseményen belül, akkor pedig abc sorrend alapján"
            elif ws.cell(sor, oszlop).value == "kicsi":
                summary3b = "A legkisebb méretre összehúzzuk oldalról a böngészőt és megnézzük hogy néz ki"
                elvart = "Minden rendben jelenik meg."
            elif ws.cell(sor, oszlop).value == "kozepes":
                summary3b = "Kb közepes méret állítjuk a böngészőt(oldalról nézve) és megnézzük, hogy hogyan néz ki."
                elvart = "Minden rendben jelenik meg."
            elif ws.cell(sor, oszlop).value == "legnagyobb":
                summary3b = "Teljes képernyőre állítjuk a böngészőt(oldalról nézve) és megnézzük, hogy hogyan néz ki."
                elvart = "Minden rendben jelenik meg."
        elif oszlop == 5:
            if ws.cell(sor, oszlop).value == "mindegy":
                summary4b = "Mindegy, hogy milyen böngészővel nézzük. Chrome, Firefox"
            elif ws.cell(sor, oszlop).value == "chrome":
                summary4b = "Chrome böngészővel nézzük."
            elif ws.cell(sor, oszlop).value == "firefox":
                summary4b = "Firefox böngészővel nézzük"
            elif ws.cell(sor, oszlop).value == "edge":
                summary4b = "Edge böngészővel nézzük"
            elif ws.cell(sor, oszlop).value == "safari":
                summary4b = "Safari böngészővel nézzük"
        elif oszlop == 6:
            if ws.cell(sor, oszlop).value == "HUN":
                summary5b = "Magyar nelvi beállításokkal nézzük. Ebben az esetben minden megjelenésnek magyar nyelv szerint helyesnek kell lennie."
                alapteszt1a = "Be kell lépni az aktuális városkártya oldalon."
                alapteszt1b = "Belépés sikerüt"
                alapteszt2a = "Be kell menni a Jegypénztár/Eladás menüpontba"
                alapteszt2b = "A belépés sikerült és betöltődéskor látszódnak az események."
            elif ws.cell(sor, oszlop).value == "EN":
                summary5b = "Angol nyelvi beállítással nézzük. Ekkor mindennek meg kell felelni az angol nyelv írásmódjának"
                alapteszt1a = "Be kell lépni az aktuális városkártya oldalon."
                alapteszt1b = "Belépés sikerüt"
                alapteszt2a = "Be kell menni a Box office/Sales menüpontba"
                alapteszt2b = "A belépés sikerült és betöltődéskor látszódnak az események."
    print(szoveg0)
    print(szoveg1)

    teljesszoveg = szoveg0 + '-' + szoveg1 + "-ITE-553"
    ws.cell(sor, eredemenyoszlop).value = teljesszoveg
    summary1 = summary1a + summary1b
    summary2 = summary2a + summary2b
    summary3 = summary3a + summary3b
    summary4 = summary4a + summary4b
    summary5 = summary5a + summary5b
    summarylista = []
    summarylista.append(summary0)
    summarylista.append(summary1)
    summarylista.append(summary2)
    summarylista.append(summary3)
    summarylista.append(summary4)
    summarylista.append(summary5)
    summary = '</br></br>'.join(summarylista)
    #print(ws.cell(sor, eredemenyoszlop).value)
    # ws1.cell(sor, eredmenysummary).value = summary
    elsosor = elsosor + 1
    # teszteset neve ami a fő sheetre kerül
    ws1.cell(elsosor,3).value = teljesszoveg
    # teszteset summary leírása
    ws1.cell(elsosor, 5).value = summary
    action0 = "Megnézzük, hogy a summary-ben leírt módon működik-e a riport."
    "Beletesszük az alap szövegeket"
    ws1.cell(elsosor, 7).value = alapteszt1a
    ws1.cell(elsosor, 8).value = alapteszt1b
    elsosor = elsosor + 1
    ws1.cell(elsosor, 7).value = alapteszt2a
    ws1.cell(elsosor, 8).value = alapteszt2b
    elsosor = elsosor + 1
    ws1.cell(elsosor,7).value = action0
    ws1.cell(elsosor, 8).value =elvart
wb.save(file2)
wb.close()
